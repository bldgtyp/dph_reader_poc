#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Phase 0 §0.5 — extract the PHPP `Areas` and `U-values` worksheets as numerical ground truth.

Read-only, `data_only=True`: openpyxl does not recalculate, so this reads the values Excel last
cached. The workbook is never written.

Cell addresses come from the `phi-rules` corpus teardowns, not from exploring the workbook:

* `rulesets/phpp-10-r1/calculators/phpp-areas/rules.md`
* `rulesets/phpp-10-r1/calculators/phpp-u-values/rules.md`

Both were verified against `PHPP_EN_V10.6_Empty.xlsx` and re-confirmed against this file.

A caveat the corpus is explicit about: **entry-block row positions vary between files** as
sections grow. `Areas` and `U-values` both use fixed-size blocks, so the addresses hold here,
but each table is located by its header text and the result reported — a silently empty
extraction would be worse than a loud failure.

Usage
-----
    uv run extract_phpp_ground_truth.py WORKBOOK.xlsm --out-dir DIR
"""

from __future__ import annotations

import argparse
import csv
import sys
import warnings
from pathlib import Path
from typing import Any, Iterator

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# openpyxl cannot parse this workbook's data-validation extension or its header/footer.
# Neither affects cell values; the warnings are noise on every load.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- Areas ---------------------------------------------------------------------------------
# Summary groups: rows 8-21 surfaces, 23-25 thermal bridges, 27 neighbour, 29 envelope total.
AREAS_SUMMARY_ROWS = list(range(8, 22)) + [23, 24, 25, 27, 29]
AREAS_SUMMARY_COLS = {
    "temp_zone": "K",
    "value": "L",
    "area_group": "M",
    "group_no": "N",
    "avg_u_value_W_m2K": "P",
    "radiation_gains_heating_kWh_a": "R",
    "radiation_load_cooling_kWh_a": "T",
}

# Surface input block. Rows 33-40 are fixed/import rows (projected footprint, TFA, the five
# window-summary imports, exterior door); user-entered surfaces run 41-141.
AREAS_SURFACE_HEADER_ROW = 32
AREAS_SURFACE_FIRST_ROW = 33
AREAS_SURFACE_LAST_ROW = 141
AREAS_SURFACE_FIRST_USER_ROW = 41
AREAS_SURFACE_COLS = {
    "area_no": "K",
    "description": "L",
    "assigned_to_group": "M",
    "quantity": "N",
    "a_m": "P",
    "b_m": "R",
    "user_defined_area_m2": "T",
    "user_defined_subtraction_m2": "V",
    "window_subtraction_m2": "X",
    "area_m2": "Z",
    "assembly_selection": "AA",
    "u_value_W_m2K": "AC",
    "azimuth": "AE",
    "inclination": "AG",
    "group_no": "BJ",
}

# --- U-values ------------------------------------------------------------------------------
# Repeating 21-row assembly blocks. Block N's header is at 7+(N-1)*21; the name/ID row is +1.
UV_FIRST_HEADER_ROW = 7
UV_BLOCK_STRIDE = 21
UV_MAX_BLOCKS = 60
UV_OFFSETS = {  # from the block header row
    "name": ("L", 1),
    "assembly_id": ("Q", 1),
    "rsi_selector": ("M", 3),
    "rse_selector": ("M", 4),
    "total_thickness_cm": ("R", 16),
    "resolved_rsi": ("M", 17),
    "resolved_rse": ("M", 18),
    "u_value_W_m2K": ("R", 18),
}
UV_LAYER_OFFSETS = (6, 14)  # layer rows: header+6 .. header+13
UV_LAYER_COLS = {
    "material": "L",
    "lambda_W_mK": "M",
    "section_2_lambda": "O",
    "section_3_lambda": "Q",
    "thickness_mm": "R",
}


def cell(ws: Worksheet, col: str, row: int) -> Any:
    return ws[f"{col}{row}"].value


def blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def areas_summary(ws: Worksheet) -> list[dict[str, Any]]:
    return [
        {"row": r} | {name: cell(ws, col, r) for name, col in AREAS_SUMMARY_COLS.items()}
        for r in AREAS_SUMMARY_ROWS
    ]


def areas_surfaces(ws: Worksheet) -> list[dict[str, Any]]:
    rows = []
    for r in range(AREAS_SURFACE_FIRST_ROW, AREAS_SURFACE_LAST_ROW + 1):
        record = {name: cell(ws, col, r) for name, col in AREAS_SURFACE_COLS.items()}
        if all(blank(v) for v in record.values()):
            continue
        record = {
            "row": r,
            "is_fixed_row": r < AREAS_SURFACE_FIRST_USER_ROW,
        } | record
        rows.append(record)
    return rows


def uvalue_blocks(ws: Worksheet) -> Iterator[tuple[int, dict[str, Any], list[dict[str, Any]]]]:
    """Walk the repeating assembly blocks, skipping unused ones.

    Raises if the last scanned block is populated: `UV_MAX_BLOCKS` is an assumption about how far
    the table runs, and silently stopping at a still-populated block would drop assemblies from the
    ground-truth extraction without any sign that it happened.
    """
    for n in range(UV_MAX_BLOCKS):
        header = UV_FIRST_HEADER_ROW + n * UV_BLOCK_STRIDE
        assembly = {
            name: cell(ws, col, header + offset) for name, (col, offset) in UV_OFFSETS.items()
        }
        layers = []
        for lr in range(header + UV_LAYER_OFFSETS[0], header + UV_LAYER_OFFSETS[1]):
            layer = {name: cell(ws, col, lr) for name, col in UV_LAYER_COLS.items()}
            if blank(layer["material"]) and blank(layer["thickness_mm"]):
                continue
            layers.append({"row": lr} | layer)
        # An assembly with no name and no layers is an unused block, not a gap in the file.
        if blank(assembly["name"]) and not layers:
            continue
        if n == UV_MAX_BLOCKS - 1:
            raise RuntimeError(
                f"U-values block {n + 1} (row {header}) is populated — the table runs past the "
                f"UV_MAX_BLOCKS={UV_MAX_BLOCKS} scan limit and assemblies beyond it would be "
                "dropped. Raise the limit and re-run."
            )
        yield header, assembly, layers


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        print(f"  !! nothing to write for {path.name}", file=sys.stderr)
        return
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    print(f"  {path.name:<34} {len(rows):5d} rows")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--out-dir", type=Path, required=True)
    args = ap.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    # Full load, ~6.6 s on the 10 MB Adelphi workbook. `read_only=True` opens in 0.7 s but makes
    # the scattered `ws["R25"]`-style lookups below pathological — the same extraction did not
    # finish in two minutes. If this ever needs to be faster, the fix is to sweep each sheet once
    # with `iter_rows(values_only=True)` into a dict, not to flip the flag. Measured 2026-08-19.
    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    for sheet in ("Areas", "U-values"):
        if sheet not in wb.sheetnames:
            print(f"worksheet {sheet!r} not in {args.workbook}", file=sys.stderr)
            return 1

    areas = wb["Areas"]
    # Locate rather than assume: if the header text has moved, the addresses below are wrong
    # and everything downstream is silently misaligned.
    header = cell(areas, "K", AREAS_SURFACE_HEADER_ROW)
    if header != "Area no.":
        print(
            f"Areas!K{AREAS_SURFACE_HEADER_ROW} is {header!r}, expected 'Area no.' — the surface "
            "table has moved; re-locate it before trusting this extraction.",
            file=sys.stderr,
        )
        return 1

    print(f"{args.workbook.name} -> {args.out_dir}/")
    write_csv(args.out_dir / "phpp_areas_summary.csv", areas_summary(areas))
    write_csv(args.out_dir / "phpp_areas_surfaces.csv", areas_surfaces(areas))

    uv = wb["U-values"]
    assemblies: list[dict[str, Any]] = []
    layers: list[dict[str, Any]] = []
    for header_row, assembly, block_layers in uvalue_blocks(uv):
        assemblies.append({"header_row": header_row, "layer_count": len(block_layers)} | assembly)
        for layer in block_layers:
            layers.append({"assembly_id": assembly["assembly_id"], "name": assembly["name"]} | layer)
    write_csv(args.out_dir / "phpp_u-values_assemblies.csv", assemblies)
    write_csv(args.out_dir / "phpp_u-values_layers.csv", layers)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
